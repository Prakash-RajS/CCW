# fastapi_app/routes/admin_dashboard.py
from fastapi import APIRouter, HTTPException, Depends, Path, Response, Cookie, Header, Query, Form, File, UploadFile
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta, date, timezone
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth, TruncWeek, TruncDay
from django.utils import timezone as django_timezone
import math
import jwt
import os
import shutil
import csv
import io
import re
import openpyxl
from fastapi.responses import StreamingResponse
from decimal import Decimal
import json
from django.contrib.auth.hashers import check_password
from django.contrib.auth.hashers import make_password
from django.db.models import Count
from fastapi_app.routes.dbconnection import ensure_db_connection, check_db_connection
# Add this near the top with other imports
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# ✅ DATABASE CONNECTION MANAGEMENT (Import from dbconnection)
from fastapi_app.services.admin_notification_service import (
    get_admin_notifications,
    create_admin_notification,
    create_notification_for_all_admins,
    clear_all_notifications  # ✅ ADD THIS
)

BASE_URL = os.getenv("BACKEND_BASE_URL", "http://67.202.26.110/api")

# ✅ IMPORT YOUR EXISTING MODELS
from creator_app.models import (
    UserData,
    AdminUser,
    UserSubscription,
    BillingHistory,
    JobPost,
    Proposal,
    WalletTransaction,
    UserPreferences,
    Contract,
    SubscriptionPlan,
    SubscriptionHistory,
    CreatorProfile,
    CollaboratorProfile,
    AdminNotificationRead,
    AdminNotification,
    UserLoginActivity
)
# ============================================================
# S3 STORAGE IMPORTS - ADD THESE
# ============================================================
from fastapi_app.routes.storage import (
    USE_S3,
    save_upload_file,
    build_full_url,
    delete_file,
    generate_presigned_url,
    ExpiryPreset,
    StoragePath,
    get_storage_path,
    get_s3_key_from_path,
)
router = APIRouter(prefix="/admin", tags=["Admin Dashboard"])

# ==============================================================================
# 🔐 CONFIGURATION
# ==============================================================================
SECRET_KEY = os.getenv("SECRET_KEY", "your_super_secret_key_123")
REFRESH_SECRET_KEY = os.getenv("REFRESH_SECRET_KEY", "your_refresh_secret_key_456")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 15  # Short-lived access token
REFRESH_TOKEN_EXPIRE_DAYS = 7  # Longer-lived refresh token

# ==============================================================================
# 🔐 MODELS
# ==============================================================================
class AdminLoginSchema(BaseModel):
    email: str
    password: str
    rememberMe: Optional[bool] = False

class TokenResponse(BaseModel):
    status: str
    message: str
    user_id: int
    name: str
    email: str
    role: str

class UserUpdateSchema(BaseModel):
    name: Optional[str] = None
    status: Optional[str] = None

class PasswordChangeSchema(BaseModel):
    new_password: str

class PreferencesSchema(BaseModel):
    theme: str
    time_zone: str
    date_format: str
    default_dashboard: str

class AdminProfileUpdateSchema(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None

class AdminPasswordChangeSchema(BaseModel):
    current_password: str
    new_password: str

# ==============================================================================
# 🔐 TOKEN FUNCTIONS
# ==============================================================================
def create_access_token(data: dict):
    """Create short-lived access token (15 minutes)"""
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire, "type": "access"})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def create_refresh_token(data: dict):
    """Create longer-lived refresh token (7 days)"""
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire, "type": "refresh"})
    return jwt.encode(to_encode, REFRESH_SECRET_KEY, algorithm=ALGORITHM)

def set_auth_cookies(response: Response, access_token: str, refresh_token: str, remember_me: bool = False):
    """Set HTTP-only cookies for authentication"""
   
    # For local development, set secure=False
    secure_cookie = False  # Set to True only in production with HTTPS
   
    # Access token - always session cookie (expires when browser closes)
    response.set_cookie(
        key="admin_access_token",
        value=access_token,
        httponly=True,
        secure=secure_cookie,
        samesite="lax",
        max_age=None,
        path="/"
    )
   
    # Refresh token - persistent cookie (long-lived if remember me)
    max_age = 30 * 24 * 60 * 60 if remember_me else None
   
    response.set_cookie(
        key="admin_refresh_token",  # CHANGED: from "refresh_token" to "admin_refresh_token"
        value=refresh_token,
        httponly=True,
        secure=secure_cookie,
        samesite="lax",
        max_age=max_age,
        path="/"
    )
# ==============================================================================
# S3 HELPER FUNCTIONS FOR ADMIN PROFILE IMAGES - ADD THIS SECTION
# ==============================================================================

async def save_admin_profile_image_s3(file: UploadFile, admin_id: int) -> str:
    """
    Save admin profile image to S3 or local storage
    S3 Folder Used: admin_profiles/
    File Type: admin
    """
    import time
    from pathlib import Path
    
    # Validate image
    allowed_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
    file_extension = Path(file.filename).suffix.lower()
    
    if file_extension not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail="Only image files are allowed"
        )
    
    # Generate unique filename
    timestamp = int(time.time())
    filename = f"admin_{admin_id}_{timestamp}{file_extension}"
    
    if USE_S3:
        # S3 path
        s3_key = f"admin_profiles/{filename}"
        await save_upload_file(file, s3_key)
        return s3_key
    else:
        # Local storage
        BASE_DIR = Path(__file__).resolve().parent.parent
        MEDIA_ROOT = BASE_DIR / "media"
        ADMIN_PROFILE_DIR = MEDIA_ROOT / "admin_profiles"
        ADMIN_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
        
        file_path = ADMIN_PROFILE_DIR / filename
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        return f"admin_profiles/{filename}"


async def delete_admin_profile_image_s3(file_path: str) -> bool:
    """
    Delete admin profile image from S3 or local storage
    """
    if not file_path:
        return False
    
    if USE_S3:
        s3_key = get_s3_key_from_path(file_path)
        return delete_file(s3_key)
    else:
        BASE_DIR = Path(__file__).resolve().parent.parent
        MEDIA_ROOT = BASE_DIR / "media"
        full_path = MEDIA_ROOT / file_path
        if full_path.exists():
            full_path.unlink()
            return True
        return False


def get_admin_profile_image_url(file_path: str, request=None) -> Optional[str]:
    """
    Get URL for admin profile image with S3 support
    """
    if not file_path:
        return None
    
    if USE_S3:
        s3_key = get_s3_key_from_path(file_path)
        return generate_presigned_url(
            s3_key=s3_key,
            expires_in=ExpiryPreset.DAILY,  # 24 hours for profile images
            force_download=False
        )
    else:
        # Local storage URL
        return f"{BASE_URL}/media/{file_path}"
    
def clear_auth_cookies(response: Response):
    """Clear authentication cookies"""
    response.delete_cookie(key="admin_access_token", path="/")  # FIXED: Changed from "access_token"
    response.delete_cookie(key="admin_refresh_token", path="/")  # CHANGED: from "refresh_token" to "admin_refresh_token"

# ==============================================================================
# 🔐 AUTH DEPENDENCY
# ==============================================================================
def get_current_admin(admin_access_token: Optional[str] = Cookie(None)):
    """
    Dependency to get current admin from access token cookie.
    Critical for all protected admin routes.
    """
    # 🔥 MOST IMPORTANT: Refresh connection BEFORE any DB operation
    ensure_db_connection()
    
    if not admin_access_token:
        raise HTTPException(
            status_code=401, 
            detail="Not authenticated"
        )
   
    try:
        payload = jwt.decode(
            admin_access_token, 
            SECRET_KEY, 
            algorithms=[ALGORITHM]
        )
       
        if payload.get("type") != "access":
            raise HTTPException(
                status_code=401, 
                detail="Invalid token type"
            )
       
        # Get admin from database
        admin = AdminUser.objects.get(
            id=payload.get("user_id"), 
            email=payload.get("sub")
        )
        return admin
        
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=401, 
            detail="Token expired"
        )
    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=401, 
            detail="Invalid token"
        )
    except AdminUser.DoesNotExist:
        raise HTTPException(
            status_code=401, 
            detail="Admin user not found"
        )
    except Exception as e:   # Catch any unexpected DB or other errors
        # Log the real error for debugging
        import logging
        logging.getLogger(__name__).error(f"get_current_admin error: {e}")
        raise HTTPException(
            status_code=500, 
            detail="Internal server error during authentication"
        )

# ==============================================================================
# 🔐 1. ADMIN LOGIN
# ==============================================================================
@router.post("/login", response_model=TokenResponse)
def admin_login(data: AdminLoginSchema, response: Response):
    """
    Login for Super Admins (AdminUser Table).
    Sets HTTP-only cookies with tokens.
    """
    # Ensure database connection
    ensure_db_connection()
   
    try:
        # Check the AdminUser table
        admin = AdminUser.objects.get(email=data.email)
        # Verify Password
        if not admin.check_password(data.password):
            # Return specific error for wrong password
            raise HTTPException(
                status_code=401,
                detail={
                    "message": "Invalid credentials",
                    "error_type": "wrong_password"
                }
            )
        # Create token payload
        token_payload = {
            "sub": admin.email,
            "user_id": admin.id,
            "name": getattr(admin, 'name', 'Admin'),
            "role": "Admin"
        }
       
        # Generate tokens
        access_token = create_access_token(token_payload)
        refresh_token = create_refresh_token({"sub": admin.email, "user_id": admin.id})
       
        # Set cookies
        set_auth_cookies(response, access_token, refresh_token, data.rememberMe)
        
        # 🔔 NOTIFICATION: Admin logged in
        create_admin_notification(
            admin=admin,
            notification_type="admin_login",
            title="Admin Login",
            subtitle=f"{admin.name or admin.email} logged into admin panel"
        )
        
        return {
            "status": "success",
            "message": "Admin login successful",
            "user_id": admin.id,
            "name": getattr(admin, 'name', 'Admin'),
            "email": admin.email,
            "role": "Admin"
        }
    except AdminUser.DoesNotExist:
        # Return specific error for user not found
        raise HTTPException(
            status_code=401,
            detail={
                "message": "Invalid credentials",
                "error_type": "user_not_found"
            }
        )

# ==============================================================================
# 🔐 2. REFRESH TOKEN ENDPOINT
# ==============================================================================
@router.post("/refresh")
def refresh_token(
    response: Response,
    admin_refresh_token: Optional[str] = Cookie(None)  # CHANGED: parameter name to match cookie
):
    """
    Refresh access token using refresh token from cookie
    """
    if not admin_refresh_token:  # CHANGED: variable name
        raise HTTPException(status_code=401, detail="No refresh token provided")
   
    try:
        # Verify refresh token
        payload = jwt.decode(admin_refresh_token, REFRESH_SECRET_KEY, algorithms=[ALGORITHM])  # CHANGED: variable name
       
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
       
        # Get user from database
        try:
            admin = AdminUser.objects.get(id=payload.get("user_id"), email=payload.get("sub"))
        except AdminUser.DoesNotExist:
            raise HTTPException(status_code=401, detail="User not found")
       
        # Create new access token
        new_access_token = create_access_token({
            "sub": admin.email,
            "user_id": admin.id,
            "name": getattr(admin, 'name', 'Admin'),
            "role": "Admin"
        })
       
        # Set new access token cookie
        response.set_cookie(
            key="admin_access_token",
            value=new_access_token,
            httponly=True,
            secure=False,  # Set to True in production
            samesite="lax",
            max_age=None,  # Session cookie
            path="/"
        )
       
        return {"status": "success", "message": "Token refreshed"}
       
    except jwt.ExpiredSignatureError:
        # Refresh token expired - clear cookies
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ==============================================================================
# 🔐 3. LOGOUT
# ==============================================================================
@router.post("/logout")
def logout(response: Response, admin: AdminUser = Depends(get_current_admin)):
    """
    Logout admin - clear auth cookies
    """
    # 🔔 NOTIFICATION: Admin logged out
    create_admin_notification(
        admin=admin,
        notification_type="admin_logout",
        title="Admin Logout",
        subtitle=f"{admin.name or admin.email} logged out from admin panel"
    )
    
    clear_auth_cookies(response)
    return {"status": "success", "message": "Logged out successfully"}

# ==============================================================================
# 🔐 4. VERIFY AUTHENTICATION (Cookie-based auth dependency)
# ==============================================================================
# get_current_admin defined above

# ==============================================================================
# 🔐 5. LEGACY ADMIN SECURITY (Header-based auth - for backward compatibility)
# ==============================================================================
def verify_admin(admin_header_id: int = Header(..., alias="user_id", description="ID of the admin making the request")):
    """
    Verifies the request is coming from a logged-in Admin.
    Checks the AdminUser table. (Legacy - uses header)
    """
    # Ensure database connection
    ensure_db_connection()
   
    try:
        # Check AdminUser table instead of UserData
        user = AdminUser.objects.get(id=admin_header_id)
        return user
    except AdminUser.DoesNotExist:
        raise HTTPException(status_code=404, detail="Admin user not found or session invalid.")

# ==============================================================================
# 🔐 6. VERIFY ENDPOINT (for frontend to check auth status)
# ==============================================================================
@router.get("/verify")
def verify_auth(current_admin: AdminUser = Depends(get_current_admin)):
    """
    Verify if the user is authenticated (cookie-based)
    """
    return {
        "status": "success",
        "authenticated": True,
        "user": {
            "id": current_admin.id,
            "email": current_admin.email,
            "name": getattr(current_admin, 'name', 'Admin')
        }
    }

# ==============================================================================
# 📊 7. DASHBOARD ENDPOINTS
# ==============================================================================
@router.get("/dashboard")
def get_dashboard(current_admin: AdminUser = Depends(get_current_admin)):
    """
    Protected route example - requires valid access token cookie
    """
    return {
        "message": "Welcome to admin dashboard",
        "admin": current_admin.email
    }

# ==============================================================================
# 📊 8. ANALYTICS PAGE ENDPOINTS
# ==============================================================================
@router.get("/overview")
def get_user_dashboard_overview(user_id: int, admin: AdminUser = Depends(get_current_admin)):
    # Ensure database connection
    ensure_db_connection()
   
    try:
        user = UserData.objects.get(id=user_id)
        has_active_plan = UserSubscription.objects.filter(
            user=user,
            plan_expires_at__gt=datetime.now()
        ).exists()
        return {
            "name": user.full_name,  # Changed from first_name to full_name
            "role": user.role,
            "show_upgrade_banner": not has_active_plan
        }
    except UserData.DoesNotExist:
        raise HTTPException(status_code=404, detail="User not found")

from django.utils import timezone as django_timezone
from datetime import timedelta

@router.get("/analytics/stats")
def get_analytics_stats(admin: AdminUser = Depends(get_current_admin)):
    """ Top 4 Cards for Analytics Page """
    # Ensure database connection
    ensure_db_connection()
   
    now = django_timezone.now()
    
    # Get first day of current month
    first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # 1. Revenue this month - FILTER BY CURRENT MONTH ONLY
    revenue_agg = SubscriptionHistory.objects.filter(
        action__in=["created", "updated"],  # only real payments
        start_date__gte=first_day_of_month,
        start_date__lte=now,
        plan_price__gt=0  # ← ADD THIS: Only count paid subscriptions
    ).aggregate(total=Sum('plan_price'))

    total_revenue = revenue_agg['total'] or 0.0  

    # 2. Active Creators & Collaborators
    creators = UserData.objects.filter(role__iexact="Creator", status__iexact="Active").count()
    collabs = UserData.objects.filter(role__iexact="Collaborator", status__iexact="Active").count()
   
    # 3. Total Subscriptions THIS MONTH (paid only, price > 0)
    subs = SubscriptionHistory.objects.filter(
        action__in=["created", "updated"],  # only real payments/subscriptions
        start_date__gte=first_day_of_month,  # Current month only
        start_date__lte=now,
        plan_price__gt=0  # Exclude free/basic plans (price = 0)
    ).values('user').distinct().count()  # ← Count unique users (one subscription per user)
    
    return {
        "revenue_month": float(total_revenue),
        "active_creators": creators,
        "active_collaborators": collabs,
        "total_subs": subs  # Now shows paid subscriptions for current month only
    }

@router.get("/analytics/user-overview")
def get_user_overview_chart(admin: AdminUser = Depends(get_current_admin)):
    """ Mixed Chart: Creators vs Collaborators vs Transactions (Last 6 Months) """
    # Ensure database connection
    ensure_db_connection()
   
    now = django_timezone.now()
    data = []
   
    for i in range(5, -1, -1):
        month_start = (now - timedelta(days=i*30)).replace(day=1)
        month_label = month_start.strftime("%b")  # Jan, Feb
       
        creators = UserData.objects.filter(role__iexact="Creator", created_at__year=month_start.year, created_at__month=month_start.month).count()
        collabs = UserData.objects.filter(role__iexact="Collaborator", created_at__year=month_start.year, created_at__month=month_start.month).count()
        txns = WalletTransaction.objects.filter(created_at__year=month_start.year, created_at__month=month_start.month).count()
        data.append({
            "Month": month_label,
            "Creator": creators,
            "Collaborator": collabs,
            "Transactions": txns
        })
    return data

@router.get("/analytics/task-performance")
def get_task_performance(admin: AdminUser = Depends(get_current_admin)):
    """
    Returns real task performance stats from the CONTRACTS table.
    """
    # Ensure database connection
    ensure_db_connection()
   
    # 1. Total Targets (Total Contracts Created)
    total_contracts = Contract.objects.count()
   
    # 2. Completed Tasks (Contracts where status is 'completed')
    completed_contracts = Contract.objects.filter(status__iexact="completed").count()
   
    # 3. Calculate "Late" vs "On Time"
    late_count = 0
    contracts = Contract.objects.filter(status__iexact="completed")
   
    for c in contracts:
        if c.end_date and c.updated_at:
            completion_date = c.updated_at.date()
            if completion_date > c.end_date:
                late_count += 1
    on_time = completed_contracts - late_count
    # 4. Growth Calculation
    now = datetime.now()
   
    if now.month == 1:
        last_month_num = 12
        last_month_year = now.year - 1
    else:
        last_month_num = now.month - 1
        last_month_year = now.year
    this_month = Contract.objects.filter(start_date__month=now.month, start_date__year=now.year).count()
    last_month = Contract.objects.filter(start_date__month=last_month_num, start_date__year=last_month_year).count()
    if last_month > 0:
        growth = ((this_month - last_month) / last_month) * 100
    else:
        growth = 100 if this_month > 0 else 0
    return {
        "total_completed": completed_contracts,
        "total_target": total_contracts,
        "on_time": on_time,
        "late": late_count,
        "tasks_this_year": Contract.objects.filter(start_date__year=now.year).count(),
        "growth": round(growth, 1)
    }

@router.get("/analytics/traffic-data")
def get_traffic_data(admin: AdminUser = Depends(get_current_admin)):
    """
    Distributes TOTAL USER COUNT into Devices/Locations using real data
    """
    ensure_db_connection()

    # DEVICE DATA from UserLoginActivity
    devices = (
        UserLoginActivity.objects
        .values("device")
        .annotate(value=Count("id"))
    )

    device_data = [
        {"name": d["device"] or "Unknown", "value": d["value"]}
        for d in devices
    ]

    # LOCATION DATA from CreatorProfile
    locations = (
        CreatorProfile.objects
        .values("location")
        .annotate(value=Count("id"))
    )

    location_data = [
        {"name": l["location"] or "Unknown", "value": l["value"]}
        for l in locations
    ]

    return {
        "device": device_data,
        "location": location_data
    }

@router.get("/analytics/revenue-splits")
def get_revenue_splits(admin: AdminUser = Depends(get_current_admin)):
    """ Pie Chart Data using fixed percentages (since no percentage fields in SubscriptionPlan) """
    ensure_db_connection()
    
    # Use fixed percentages since your SubscriptionPlan doesn't have percentage fields
    # You can adjust these values as needed
    platform_percentage = 10
    creator_percentage = 60
    collaborator_percentage = 30
    
    # Get total revenue from SubscriptionHistory (more accurate)
    total_revenue = SubscriptionHistory.objects.filter(
        action__in=["created", "updated"]
    ).aggregate(total=Sum('plan_price'))['total'] or 0.0
    
    # If no revenue from subscriptions, use BillingHistory as fallback
    if total_revenue == 0:
        total_revenue = BillingHistory.objects.filter(
            status="paid"
        ).aggregate(total=Sum('amount'))['total'] or 0.0
    
    # Calculate split values based on total revenue
    platform_total = (total_revenue * platform_percentage) / 100
    creator_total = (total_revenue * creator_percentage) / 100
    collaborator_total = (total_revenue * collaborator_percentage) / 100
    
    total = total_revenue if total_revenue > 0 else 1
    
    return {
        "splits": [
            {
                "name": "Platform Fees",
                "value": float(platform_total),
                "percentage": round((platform_total / total) * 100, 2),
                "color": "#DC2626"  # Dark Red
            },
            {
                "name": "Creator",
                "value": float(creator_total),
                "percentage": round((creator_total / total) * 100, 2),
                "color": "#7C3AED"  # Dark Purple
            },
            {
                "name": "Collaborator",
                "value": float(collaborator_total),
                "percentage": round((collaborator_total / total) * 100, 2),
                "color": "#059669"  # Dark Green
            }
        ]
    }

@router.get("/analytics/top-collaborators")
def get_top_collaborators(limit: int = 5, admin: AdminUser = Depends(get_current_admin)):
    """ Top Collaborator List ranked by wallet balance """
    ensure_db_connection()
    
    # ✅ Import S3 functions
    from fastapi_app.routes.storage import (
        generate_presigned_url_with_cache, 
        get_s3_key_from_path, 
        USE_S3
    )
    
    top_users = UserData.objects.filter(
        role__iexact="collaborator"
    ).order_by('-wallet__balance')[:limit]
    
    results = []
    
    for idx, u in enumerate(top_users, 1):
        wallet_bal = u.wallet.balance if hasattr(u, 'wallet') else 0.0
        
        profile_image = None
        
        # ✅ Get profile image with S3 support
        if u.profile_picture:
            profile_pic = str(u.profile_picture)
            if profile_pic.startswith(('http://', 'https://')):
                profile_image = profile_pic
            elif USE_S3:
                s3_key = get_s3_key_from_path(profile_pic)
                if s3_key:
                    profile_image = generate_presigned_url_with_cache(
                        s3_key=s3_key,
                        expires_in=86400
                    )
            else:
                profile_image = f"{BASE_URL}/media/{profile_pic}"
        
        # ✅ Fallback to CollaboratorProfile
        if not profile_image:
            try:
                collab_profile = CollaboratorProfile.objects.filter(user=u).first()
                if collab_profile and collab_profile.profile_picture:
                    profile_pic = str(collab_profile.profile_picture)
                    if USE_S3:
                        s3_key = get_s3_key_from_path(profile_pic)
                        if s3_key:
                            profile_image = generate_presigned_url_with_cache(
                                s3_key=s3_key,
                                expires_in=86400
                            )
                    else:
                        profile_image = f"{BASE_URL}/media/{profile_pic}"
            except:
                pass
        
        # Get contracts
        contracts = Contract.objects.filter(collaborator=u)
        total_jobs = contracts.count()
        completed_jobs = contracts.filter(status__iexact="completed").count()
        completion_rate = round((completed_jobs / total_jobs * 100), 1) if total_jobs > 0 else 0
        total_budget = contracts.aggregate(total=Sum('budget'))['total'] or 0.0
        
        completed_contracts = contracts.filter(status__iexact="completed")
        total_earnings = completed_contracts.aggregate(total=Sum('budget'))['total'] or 0.0
        avg_job_value = round(total_earnings / completed_jobs, 2) if completed_jobs > 0 else 0
        
        thirty_days_ago = django_timezone.now() - timedelta(days=30)
        monthly_revenue = completed_contracts.filter(
            updated_at__gte=thirty_days_ago
        ).aggregate(total=Sum('budget'))['total'] or 0.0
        
        results.append({
            "rank": idx,
            "name": u.full_name,
            "email": u.email,
            "earnings": float(wallet_bal),
            "total_jobs": total_jobs,
            "completed_jobs": completed_jobs,
            "completion_rate": completion_rate,
            "avg_job_value": avg_job_value,
            "monthly_revenue": float(monthly_revenue),
            "total_budget": float(total_budget),
            "joined_date": u.created_at.strftime("%d %b %Y"),
            "profile_image": profile_image,
            "status": u.status
        })
    
    return results
# ==============================================================================
# 📊 9. DASHBOARD OVERVIEW (Top Cards & Charts)
# ==============================================================================
@router.get("/dashboard/stats")
def get_dashboard_stats(admin: AdminUser = Depends(get_current_admin)):
    # Ensure database connection
    ensure_db_connection()
   
    total_users = UserData.objects.count()
    # Count active projects: in_progress, awaiting, pending, in_review
    active_projects = Contract.objects.filter(
        Q(status__iexact="in_progress") | 
        Q(status__iexact="awaiting") | 
        Q(status__iexact="pending") |
        Q(status__iexact="in_review")
    ).count()
    completed_tasks = Contract.objects.filter(status__iexact="completed").count()
    revenue_agg = SubscriptionHistory.objects.filter(
        action__in=["created", "updated"]
    ).aggregate(total=Sum('plan_price'))

    total_revenue = float(revenue_agg['total'] or 0)
    return {
        "admin_name": getattr(admin, 'name', 'Admin'),
        "total_users": total_users,
        "active_projects": active_projects,
        "completed_tasks": completed_tasks,
        "total_revenue": float(total_revenue)
    }

@router.get("/dashboard/charts/revenue")
def get_revenue_chart(
    filter: str = Query("Yearly", enum=["Weekly", "Monthly", "Yearly"]),
    year: int = Query(default=django_timezone.now().year, description="Select year for Yearly view"),
    admin: AdminUser = Depends(get_current_admin)
):
    # Ensure database connection
    ensure_db_connection()
   
    now = django_timezone.now()
    labels = []
    data = []
    if filter == "Weekly":
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            labels.append(day.strftime("%a"))
            total = SubscriptionHistory.objects.filter(
                start_date__date=day.date(),
                action__in=["created", "updated"]
            ).aggregate(s=Sum('plan_price'))['s'] or 0            
            data.append(float(total))
           
    elif filter == "Monthly":
        for i in range(3, -1, -1):
            week_start = now - timedelta(weeks=i)
            labels.append(f"Week {4-i}")
            total = SubscriptionHistory.objects.filter(
                start_date__week=week_start.isocalendar()[1],
                start_date__year=week_start.year,
                action__in=["created", "updated"]
            ).aggregate(s=Sum('plan_price'))['s'] or 0            
            data.append(float(total))
    else:  # Yearly
        labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        for i in range(1, 13):
            total = SubscriptionHistory.objects.filter(
                start_date__month=i,
                start_date__year=year,
                action__in=["created", "updated"]
            ).aggregate(s=Sum('plan_price'))['s'] or 0
            data.append(float(total))
    return {
        "labels": labels,
        "data": data,
        "growth_percentage": 0
    }

@router.get("/dashboard/charts/project-status")
def get_project_status_charts(
    time_range: str = "all",
    admin: AdminUser = Depends(get_current_admin)
):
    """
    Get project status distribution percentages for dashboard charts.
    """
    ensure_db_connection()

    try:
        now = django_timezone.now()
        base_query = Contract.objects.all()

        # Apply time range filters
        if time_range == "today":
            base_query = base_query.filter(updated_at__date=now.date())

        elif time_range == "yesterday":
            base_query = base_query.filter(
                updated_at__date=now.date() - timedelta(days=1)
            )

        elif time_range == "week":
            base_query = base_query.filter(
                updated_at__gte=now - timedelta(days=7)
            )

        elif time_range == "month":
            base_query = base_query.filter(
                updated_at__gte=now - timedelta(days=30)
            )

        # Total contracts
        total = base_query.count()

        # Completed projects
        completed = base_query.filter(
            status__iexact="completed"
        ).count()

        # On Hold projects
        on_hold = base_query.filter(
            Q(status__iexact="pending") |
            Q(status__iexact="awaiting")
        ).count()

        # Everything else considered In Progress
        in_progress = total - completed - on_hold

        # Calculate percentages
        if total > 0:
            completed_pct = round((completed / total) * 100)
            on_hold_pct = round((on_hold / total) * 100)

            # Remaining percentage automatically goes to in_progress
            in_progress_pct = 100 - completed_pct - on_hold_pct
        else:
            completed_pct = 0
            on_hold_pct = 0
            in_progress_pct = 0

        return {
            "completed": completed_pct,
            "on_hold": on_hold_pct,
            "in_progress": in_progress_pct,
            "total": total,
            "completed_count": completed,
            "on_hold_count": on_hold,
            "in_progress_count": in_progress,
            "time_range": time_range
        }

    except Exception as e:
        import logging
        logging.getLogger(__name__).error(
            f"Error in get_project_status_charts: {e}"
        )

        raise HTTPException(
            status_code=500,
            detail="Failed to fetch project status data"
        )
        
        
@router.get("/dashboard/charts/progress")
def get_progress_chart(
    filter: str = Query("Week", enum=["Week", "Month", "Year"]),
    admin: AdminUser = Depends(get_current_admin)
):
    # Ensure database connection
    ensure_db_connection()
   
    now = django_timezone.now()
    labels = []
    values = []
    if filter == "Week":
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            labels.append(day.strftime("%d %b"))
            count = JobPost.objects.filter(created_at__date=day.date()).count()
            values.append(count)
    elif filter == "Month":
        for i in range(5, -1, -1):
            day = now - timedelta(days=i*5)
            labels.append(day.strftime("%d %b"))
            count = JobPost.objects.filter(
                created_at__date__lte=day.date(),
                created_at__date__gt=(day-timedelta(days=5)).date()
            ).count()
            values.append(count)
    elif filter == "Year":
        labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        for i in range(1, 13):
            count = JobPost.objects.filter(
                created_at__year=now.year,
                created_at__month=i
            ).count()
            values.append(count)
    
    # Count ALL contracts properly
    total_contracts = Contract.objects.count()
    # Count completed contracts
    completed_contracts = Contract.objects.filter(status__iexact="completed").count()
    # Count in_progress contracts (including all active statuses)
    in_progress_contracts = Contract.objects.filter(
        Q(status__iexact="in_progress") | 
        Q(status__iexact="awaiting") | 
        Q(status__iexact="pending") |
        Q(status__iexact="in_review")
    ).count()
    
    if total_contracts > 0:
        growth_percentage = round(
            (completed_contracts / total_contracts) * 100
        )
    else:
        growth_percentage = 0

    return {
        "labels": labels,
        "data": values,
        "all_task": total_contracts,
        "done": completed_contracts,
        "in_progress": in_progress_contracts,
        "growth_percentage": growth_percentage
    }

@router.get("/dashboard/active-projects")
def get_active_projects_table(
    limit: int = 100,
    search: Optional[str] = Query(None, description="Search by client name or project title"),
    admin: AdminUser = Depends(get_current_admin)
):
    """
    Active Projects Table (Source: Contracts) - Supports Milestone Projects
    Returns unique projects with milestone progress
    """
    ensure_db_connection()
    
    # ✅ Import S3 functions
    from fastapi_app.routes.storage import (
        generate_presigned_url_with_cache, 
        get_s3_key_from_path, 
        USE_S3
    )

    # Get ALL contracts that are not completed or cancelled
    base_query = Contract.objects.filter(
        ~Q(status__iexact="completed") &
        ~Q(status__iexact="cancelled")
    ).select_related('creator', 'job', 'collaborator').order_by('-updated_at')

    # Apply search filter if provided
    if search:
        base_query = base_query.filter(
            Q(creator__full_name__icontains=search) |
            Q(job__title__icontains=search) |
            Q(description__icontains=search) |
            Q(collaborator__full_name__icontains=search)
        )

    projects = base_query[:limit] if limit > 0 else base_query

    data = []
    now = django_timezone.now()

    # Track seen projects to prevent duplicates
    seen_projects = set()

    for p in projects:
        # Create unique key (client_id + project_title)
        client_id = p.creator.id if p.creator else None
        project_title = p.job.title if p.job else p.description

        unique_key = (client_id, project_title)

        # Skip duplicates
        if unique_key in seen_projects:
            continue
        seen_projects.add(unique_key)

        # Get client info
        client_name = "Unknown"
        client_profile_image = None
        project_owner_type = "Creator"
        project_owner_name = ""

        if p.collaborator:
            project_owner_name = p.collaborator.full_name or p.collaborator.email.split('@')[0]
            project_owner_type = "Collaborator"

        if p.creator:
            client_name = p.creator.full_name or p.creator.email.split('@')[0]
            
            # ✅ Get client profile image with S3 support
            if p.creator.profile_picture:
                profile_pic = str(p.creator.profile_picture)
                if profile_pic.startswith(('http://', 'https://')):
                    client_profile_image = profile_pic
                elif USE_S3:
                    s3_key = get_s3_key_from_path(profile_pic)
                    if s3_key:
                        client_profile_image = generate_presigned_url_with_cache(
                            s3_key=s3_key,
                            expires_in=86400
                        )
                else:
                    client_profile_image = f"{BASE_URL}/media/{profile_pic}"
            
            # Fallback to CreatorProfile if no profile picture in UserData
            if not client_profile_image and p.creator.role and p.creator.role.lower() == "creator":
                try:
                    creator_profile = CreatorProfile.objects.filter(user=p.creator).first()
                    if creator_profile and creator_profile.profile_picture:
                        profile_pic = str(creator_profile.profile_picture)
                        if USE_S3:
                            s3_key = get_s3_key_from_path(profile_pic)
                            if s3_key:
                                client_profile_image = generate_presigned_url_with_cache(
                                    s3_key=s3_key,
                                    expires_in=86400
                                )
                        else:
                            client_profile_image = f"{BASE_URL}/media/{profile_pic}"
                except:
                    pass

        # Check if milestone-based
        is_milestone_based = False
        milestones_data = []
        completed_milestones = 0
        total_milestones = 0
        progress_percent = 0
        next_milestone_due = None
        next_milestone_description = ""

        if hasattr(p, 'milestones_data') and p.milestones_data and len(p.milestones_data) > 0:
            is_milestone_based = True
            milestones_data = p.milestones_data
            total_milestones = len(milestones_data)
            completed_milestones = sum(1 for m in milestones_data if m.get('status') == 'paid')

            if total_milestones > 0:
                progress_percent = int((completed_milestones / total_milestones) * 100)

            for milestone in milestones_data:
                if milestone.get('status') in ['in_progress', 'pending', 'submitted']:
                    if milestone.get('due_date'):
                        next_milestone_due = milestone.get('due_date')
                        next_milestone_description = milestone.get('description', '')
                    break

        # Calculate timeline
        start_date_str = "Not set"
        end_date_str = "Not set"
        duration_str = "Not started"
        days_elapsed = 0
        days_remaining = 0
        total_days = 0

        if p.start_date:
            start_date_str = p.start_date.strftime("%d %b %Y")
        if p.end_date:
            end_date_str = p.end_date.strftime("%d %b %Y")

        if is_milestone_based and next_milestone_due:
            try:
                milestone_due = datetime.strptime(next_milestone_due, "%Y-%m-%d").date()
                today = now.date()

                if milestone_due < today:
                    days_overdue = (today - milestone_due).days
                    duration_str = f"⚠️ Milestone '{next_milestone_description}' overdue by {days_overdue} days"
                else:
                    days_until = (milestone_due - today).days
                    duration_str = f"📌 Milestone '{next_milestone_description}' due in {days_until} days"

                if completed_milestones > 0:
                    duration_str = f"{completed_milestones}/{total_milestones} milestones • {duration_str}"
            except:
                duration_str = f"Milestone {completed_milestones}/{total_milestones} completed"
        else:
            if p.start_date and p.end_date:
                total_days = (p.end_date - p.start_date).days
                if total_days > 0:
                    today = now.date()
                    if today < p.start_date:
                        duration_str = f"Starts {p.start_date.strftime('%d %b')}"
                        days_remaining = total_days
                    elif today >= p.end_date:
                        if p.status != "completed":
                            duration_str = f"Overdue by {(today - p.end_date).days} days"
                        else:
                            duration_str = "Completed"
                        days_elapsed = total_days
                        days_remaining = 0
                    else:
                        days_elapsed = (today - p.start_date).days
                        days_remaining = (p.end_date - today).days
                        if not is_milestone_based:
                            progress_percent = int((days_elapsed / total_days) * 100)
                        duration_str = f"Day {days_elapsed} of {total_days}"

        # Format status
        contract_status = p.status.replace('_', ' ').title() if p.status else "Pending"
        if is_milestone_based:
            if completed_milestones == total_milestones:
                contract_status = "All Milestones Completed"
            elif p.status == "in_review":
                contract_status = f"Milestone {completed_milestones + 1} Under Review"
            else:
                contract_status = f"{completed_milestones}/{total_milestones} Milestones Done"

        project_category = "General"
        if is_milestone_based:
            project_category = "Milestone Based"
        elif p.job:
            if hasattr(p.job, 'timeline') and p.job.timeline:
                project_category = p.job.timeline.capitalize()
            elif hasattr(p.job, 'expertise_level') and p.job.expertise_level:
                project_category = p.job.expertise_level.capitalize()

        data.append({
            "client_name": client_name,
            "client_id": client_id,
            "client_profile_image": client_profile_image,
            "project_title": project_title,
            "price": float(p.budget) if p.budget else 0.0,
            "delivered_in": duration_str,
            "start_date": start_date_str,
            "end_date": end_date_str,
            "progress": progress_percent,
            "status": contract_status,
            "project_owner": project_owner_name,
            "project_owner_type": project_owner_type,
            "project_category": project_category,
            "is_milestone_based": is_milestone_based,
            "completed_milestones": completed_milestones,
            "total_milestones": total_milestones,
            "milestones_data": milestones_data,
            "next_milestone_due": next_milestone_due,
            "next_milestone_description": next_milestone_description
        })

    return data

# ==============================================================================
# 👥 10. USER MANAGEMENT (CRUD)
# ==============================================================================
@router.get("/users")
def get_all_users(
    role: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
    admin: AdminUser = Depends(get_current_admin)
):
    ensure_db_connection()
   
    query = UserData.objects.all().order_by('-created_at')
    if role:
        query = query.filter(role__iexact=role)
    if status:
        query = query.filter(status__iexact=status)
    if search:
        query = query.filter(
            Q(full_name__icontains=search) |
            Q(email__icontains=search)
        )
    total = query.count()
    start = (page - 1) * page_size
    end = start + page_size
    users = query[start:end]
    results = []
    
    # ✅ Import S3 functions
    from fastapi_app.routes.storage import (
        generate_presigned_url_with_cache, 
        get_s3_key_from_path, 
        USE_S3
    )
    
    for u in users:
        user_status = getattr(u, 'status', 'Active')
        profile_image = None
        
        # ✅ Check if user has profile picture in UserData
        if u.profile_picture:
            profile_pic = str(u.profile_picture)
            if profile_pic.startswith(('http://', 'https://')):
                profile_image = profile_pic
            elif USE_S3:
                # ✅ Generate S3 presigned URL
                s3_key = get_s3_key_from_path(profile_pic)
                if s3_key:
                    profile_image = generate_presigned_url_with_cache(
                        s3_key=s3_key,
                        expires_in=86400  # 24 hours
                    )
            else:
                profile_image = f"{BASE_URL}/media/{profile_pic}"
        
        # ✅ Fallback to role-based profile (CreatorProfile or CollaboratorProfile)
        if not profile_image:
            try:
                if u.role and u.role.lower() == "creator":
                    creator_profile = CreatorProfile.objects.filter(user=u).first()
                    if creator_profile and creator_profile.profile_picture:
                        profile_pic = str(creator_profile.profile_picture)
                        if USE_S3:
                            s3_key = get_s3_key_from_path(profile_pic)
                            if s3_key:
                                profile_image = generate_presigned_url_with_cache(
                                    s3_key=s3_key,
                                    expires_in=86400
                                )
                        else:
                            profile_image = f"{BASE_URL}/media/{profile_pic}"
                            
                elif u.role and u.role.lower() == "collaborator":
                    collab_profile = CollaboratorProfile.objects.filter(user=u).first()
                    if collab_profile and collab_profile.profile_picture:
                        profile_pic = str(collab_profile.profile_picture)
                        if USE_S3:
                            s3_key = get_s3_key_from_path(profile_pic)
                            if s3_key:
                                profile_image = generate_presigned_url_with_cache(
                                    s3_key=s3_key,
                                    expires_in=86400
                                )
                        else:
                            profile_image = f"{BASE_URL}/media/{profile_pic}"
            except Exception as e:
                logger.warning(f"Error getting role profile for user {u.id}: {e}")
        
        results.append({
            "id": u.id,
            "full_name": u.full_name or "",
            "email": u.email,
            "role": u.role,
            "status": user_status,
            "joined_date": u.created_at.strftime("%B %d, %Y"),
            "last_active": "Recently",
            "profile_image": profile_image
        })
    
    return {
        "total_users": total,
        "page": page,
        "page_size": page_size,
        "data": results
    }
# ==============================================================================
# CREATE USER
# ==============================================================================
@router.post("/users")
def create_user(
    full_name: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
    password: str = Form(...),
    status: Optional[str] = Form("Active"),
    admin: AdminUser = Depends(get_current_admin)
):
    ensure_db_connection()
    # 🔴 EMAIL VALIDATION
    allowed_domains = ["gmail.com", "yahoo.com", "outlook.com"]
    domain = email.split("@")[-1]
    if domain not in allowed_domains:
        raise HTTPException(status_code=400, detail="Invalid email domain")
    # 🔴 PASSWORD VALIDATION
    password_regex = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&]).{8,}$'
    if not re.match(password_regex, password):
        raise HTTPException(
            status_code=400,
            detail="Weak password (min 8 chars, uppercase, lowercase, number, special char)"
        )
    # 🔴 CHECK DUPLICATE EMAIL
    if UserData.objects.filter(email=email).exists():
        raise HTTPException(status_code=400, detail="Email already registered")
    # 🔐 HASH PASSWORD
    hashed_password = make_password(password)
    # ✅ CREATE USER
    new_user = UserData.objects.create(
        full_name=full_name,
        email=email,
        role=role,
        password=hashed_password,
        status=status
    )
    
    # 🔔 NOTIFICATION: New user created by admin
    create_notification_for_all_admins(
        notification_type="user_created",
        title="New User Created",
        subtitle=f"Admin {admin.name or admin.email} created user: {full_name} ({email})",
        exclude_admin=None
    )
    
    return {
        "status": "success",
        "message": "User created successfully",
        "user_id": new_user.id
    }

# ==============================================================================
# UPDATE USER (FIXED)
# ==============================================================================
@router.put("/users/{user_id}")
def update_user(
    user_id: int,
    data: UserUpdateSchema,
    admin: AdminUser = Depends(get_current_admin)
):
    ensure_db_connection()
   
    try:
        user = UserData.objects.get(id=user_id)
        old_status = user.status
        old_name = user.full_name
        
        # ✅ UPDATE NAME - Now using full_name directly
        if data.name:
            user.full_name = data.name.strip()
        # ✅ UPDATE STATUS (FIXED)
        if data.status:
            user.status = data.status
        user.save()
        
        # 🔔 NOTIFICATION: User updated by admin
        changes = []
        if data.name and data.name != old_name:
            changes.append(f"name changed from '{old_name}' to '{data.name}'")
        if data.status and data.status != old_status:
            changes.append(f"status changed from '{old_status}' to '{data.status}'")
        
        if changes:
            create_notification_for_all_admins(
                notification_type="user_updated",
                title="User Updated",
                subtitle=f"Admin {admin.name or admin.email} updated user {user.full_name}: {', '.join(changes)}",
                exclude_admin=None
            )
        
        return {
            "status": "success",
            "message": "User updated successfully"
        }
    except UserData.DoesNotExist:
        raise HTTPException(status_code=404, detail="User not found")

# ==============================================================================
# DELETE USER
# ==============================================================================
@router.delete("/users/{user_id}")
def delete_user(user_id: int, admin: AdminUser = Depends(get_current_admin)):
    ensure_db_connection()
   
    try:
        user = UserData.objects.get(id=user_id)
        user_name = user.full_name
        user_email = user.email
        user.delete()
        
        # 🔔 NOTIFICATION: User deleted by admin
        create_notification_for_all_admins(
            notification_type="user_deleted",
            title="User Deleted",
            subtitle=f"Admin {admin.name or admin.email} deleted user: {user_name} ({user_email})",
            exclude_admin=None
        )
        
        return {
            "status": "success",
            "message": "User deleted"
        }
    except UserData.DoesNotExist:
        raise HTTPException(status_code=404, detail="User not found")
   
# ==============================================================================
# 💳 11. SUBSCRIPTION MANAGEMENT
# ==============================================================================
@router.get("/subscriptions/stats")
def get_subscription_stats(admin: AdminUser = Depends(get_current_admin)):
    """Get subscription statistics for dashboard cards - DYNAMIC"""
    ensure_db_connection()
   
    try:
        from django.db.models import Q
        from collections import defaultdict
        now = django_timezone.now()
        
        # Get ALL active subscriptions
        active_subscriptions = UserSubscription.objects.filter(
            status__in=['active', 'trialing'],
            plan_end_date__gt=now
        ).select_related('user')
        
        # Get unique user IDs with active subscriptions
        subscriber_ids = list(active_subscriptions.values_list('user_id', flat=True).distinct())
        total_subscribers = len(subscriber_ids)
        
        # Get all unique plan names from active subscriptions
        plan_stats = defaultdict(lambda: {"creator": 0, "collaborator": 0, "total": 0})
        
        for sub in active_subscriptions:
            # Get plan name (prefer current_plan, fallback to plan_name)
            plan_name = sub.current_plan or sub.plan_name or "Unknown"
            
            # Get user role
            role = sub.user.role.lower() if sub.user and sub.user.role else "unknown"
            
            if role == "creator":
                plan_stats[plan_name]["creator"] += 1
            elif role == "collaborator":
                plan_stats[plan_name]["collaborator"] += 1
            
            plan_stats[plan_name]["total"] += 1
        
        # Get ALL active plans from SubscriptionPlan model (including those with 0 subscribers)
        all_plans = SubscriptionPlan.objects.filter(is_active=True)
        
        # Ensure all active plans are included even if they have 0 subscribers
        for plan in all_plans:
            plan_name = plan.name
            if plan_name not in plan_stats:
                plan_stats[plan_name] = {"creator": 0, "collaborator": 0, "total": 0}
        
        # Users with NO active subscription
        all_user_ids = set(UserData.objects.values_list('id', flat=True))
        users_without_subs = len(all_user_ids - set(subscriber_ids))
        
        # Prepare response with role information
        plans_data = []
        for plan in all_plans:
            plan_name = plan.name
            counts = plan_stats.get(plan_name, {"creator": 0, "collaborator": 0, "total": 0})
            plans_data.append({
                "name": plan_name,
                "creator_count": counts["creator"],
                "collaborator_count": counts["collaborator"],
                "total_count": counts["total"],
                "role": plan.role  # 'creator', 'collaborator', or 'both'
            })
        
        # print(f"📊 Dynamic Stats Summary:")
        # print(f"  - Total Subscribers: {total_subscribers}")
        # print(f"  - Plans found: {len(plans_data)}")
        # for plan in plans_data:
        #     print(f"    - {plan['name']} (role: {plan['role']}): Creator={plan['creator_count']}, Collaborator={plan['collaborator_count']}")
        # print(f"  - Users without subscription: {users_without_subs}")
       
        return {
            "total_subscribers": total_subscribers,
            "users_without_subscription": users_without_subs,
            "plans": plans_data  # Dynamic plans data with role info
        }
        
    except Exception as e:
        # print(f"❌ Error in get_subscription_stats: {e}")
        import traceback
        traceback.print_exc()
        return {
            "total_subscribers": 0,
            "users_without_subscription": 0,
            "plans": []
        }
    
@router.get("/subscriptions/plans")
def get_subscription_plans(admin: AdminUser = Depends(get_current_admin)):
    """Get all subscription plans with user counts"""
    # Ensure database connection
    ensure_db_connection()
   
    try:
        plans = SubscriptionPlan.objects.filter(is_active=True).order_by('price')
        result = []
       
        for plan in plans:
            # Count active users on this plan
            user_count = UserSubscription.objects.filter(
                Q(current_plan__iexact=plan.name) | Q(plan_name__iexact=plan.name),
                status__in=['active', 'trialing']
            ).count()
           
            # Get features - ensure it's a list
            features = plan.features
            if isinstance(features, str):
                try:
                    features = json.loads(features)
                except:
                    features = []
           
            # Calculate discounted price
            discounted_price = float(plan.price)
            if plan.discount_percentage and plan.discount_percentage > 0:
                discount = (plan.discount_percentage / 100) * float(plan.price)
                discounted_price = float(plan.price) - discount
           
            # Determine yearly price display
            yearly_price = ""
            if plan.duration == "monthly":
                yearly_total = float(plan.price) * 12
                if plan.discount_percentage and plan.discount_percentage > 0:
                    yearly_total = discounted_price * 12
                yearly_price = f"${yearly_total:.0f} / Year"
            elif plan.duration == "yearly":
                yearly_price = f"${float(plan.price):.0f} / Year"
           
            plan_data = {
                "id": plan.id,
                "name": plan.name,
                "price": f"${float(plan.price):.0f}/month" if plan.duration == "monthly" else f"${float(plan.price):.0f}",
                "yearly": yearly_price,
                "users": str(user_count),
                "features": [f.get('title', f) if isinstance(f, dict) else str(f) for f in features[:4]],
                "duration": plan.duration,
                "role": getattr(plan, 'role', 'both'),
                "discount_code": plan.discount_code,
                "discount_percentage": plan.discount_percentage,
                "discounted_price": round(discounted_price, 2)
            }
            result.append(plan_data)
       
        return {"plans": result}
       
    except Exception as e:
        # print(f"Error in get_subscription_plans: {e}")
        return {"plans": []}

@router.get("/subscriptions/history")
def get_subscription_history(
    limit: int = 20,
    search: Optional[str] = None,
    admin: AdminUser = Depends(get_current_admin)
):
    """Get subscription history from SubscriptionHistory table with user details and profile images"""
    ensure_db_connection()
   
    try:
        query = SubscriptionHistory.objects.select_related('user').all().order_by('-created_at')
       
        if search:
            query = query.filter(
                Q(user__email__icontains=search) |
                Q(user__full_name__icontains=search) |
                Q(plan_name__icontains=search)
            )
       
        history_records = query[:limit]
        result = []
        
        # ✅ Import S3 functions
        from fastapi_app.routes.storage import (
            generate_presigned_url_with_cache, 
            get_s3_key_from_path, 
            USE_S3
        )
       
        for idx, history in enumerate(history_records):
            user = history.user
           
            if not user:
                continue
           
            profile_image = None
            avatar_color = "#8B5CF6"  # Default purple
           
            # Set avatar color based on role
            if user.role and user.role.lower() == "creator":
                avatar_color = "#8B5CF6"
            elif user.role and user.role.lower() == "collaborator":
                avatar_color = "#10B981"
           
            # ✅ Check UserData profile picture first
            if user.profile_picture:
                try:
                    profile_pic = str(user.profile_picture)
                    if USE_S3:
                        s3_key = get_s3_key_from_path(profile_pic)
                        if s3_key:
                            profile_image = generate_presigned_url_with_cache(
                                s3_key=s3_key,
                                expires_in=86400
                            )
                    else:
                        profile_image = f"{BASE_URL}/media/{profile_pic}"
                except Exception as e:
                    logger.warning(f"Error getting UserData profile for user {user.id}: {e}")
           
            # ✅ Fallback to CreatorProfile or CollaboratorProfile
            if not profile_image:
                if user.role and user.role.lower() == "creator":
                    try:
                        creator_profile = CreatorProfile.objects.filter(user=user).first()
                        if creator_profile and creator_profile.profile_picture:
                            profile_pic = str(creator_profile.profile_picture)
                            if USE_S3:
                                s3_key = get_s3_key_from_path(profile_pic)
                                if s3_key:
                                    profile_image = generate_presigned_url_with_cache(
                                        s3_key=s3_key,
                                        expires_in=86400
                                    )
                            else:
                                profile_image = f"{BASE_URL}/media/{profile_pic}"
                    except Exception as e:
                        logger.warning(f"Error getting CreatorProfile for user {user.id}: {e}")

                elif user.role and user.role.lower() == "collaborator":
                    try:
                        collab_profile = CollaboratorProfile.objects.filter(user=user).first()
                        if collab_profile and collab_profile.profile_picture:
                            profile_pic = str(collab_profile.profile_picture)
                            if USE_S3:
                                s3_key = get_s3_key_from_path(profile_pic)
                                if s3_key:
                                    profile_image = generate_presigned_url_with_cache(
                                        s3_key=s3_key,
                                        expires_in=86400
                                    )
                            else:
                                profile_image = f"{BASE_URL}/media/{profile_pic}"
                    except Exception as e:
                        logger.warning(f"Error getting CollaboratorProfile for user {user.id}: {e}")

            # Get username
            username = user.email.split('@')[0] if user.email else "user"
           
            if user.role and user.role.lower() == "creator":
                try:
                    creator_profile = CreatorProfile.objects.filter(user=user).first()
                    if creator_profile and creator_profile.creator_name:
                        username = creator_profile.creator_name
                except Exception:
                    pass
            elif user.role and user.role.lower() == "collaborator":
                try:
                    collab_profile = CollaboratorProfile.objects.filter(user=user).first()
                    if collab_profile and collab_profile.name:
                        username = collab_profile.name
                except Exception:
                    pass

            full_name = user.full_name or ""
            if not full_name:
                full_name = user.email.split('@')[0] if user.email else "User"
           
            date_str = history.start_date.strftime("%B %d, %Y") if history.start_date else "N/A"
           
            result.append({
                "id": history.id,
                "s_no": idx + 1,
                "full_name": full_name,
                "email": user.email,
                "username": username,
                "role": user.role or "Creator",
                "date": date_str,
                "plan": history.plan_name or "Free",
                "plan_price": float(history.plan_price) if history.plan_price else 0,
                "duration": history.duration,
                "start_date": history.start_date.isoformat() if history.start_date else None,
                "end_date": history.end_date.isoformat() if history.end_date else None,
                "status": history.status,
                "action": history.action,
                "invoice_number": history.invoice_number,
                "stripe_subscription_id": history.stripe_subscription_id,
                "profile_image": profile_image,
                "avatar_color": avatar_color
            })
       
        return {"history": result}
       
    except Exception as e:
        logger.error(f"Error in get_subscription_history: {e}")
        import traceback
        traceback.print_exc()
        return {"history": []}

# ==============================================================================
# ⚙️ 12. SETTINGS
# ==============================================================================
@router.put("/users/{user_id}/change-password")
def change_user_password(user_id: int, data: PasswordChangeSchema, admin: AdminUser = Depends(get_current_admin)):
    # Ensure database connection
    ensure_db_connection()
   
    try:
        user = UserData.objects.get(id=user_id)
        user.password = make_password(data.new_password)
        user.save()
        
        # 🔔 NOTIFICATION: Admin changed user password
        create_notification_for_all_admins(
            notification_type="user_password_changed",
            title="User Password Changed",
            subtitle=f"Admin {admin.name or admin.email} changed password for user: {user.full_name} ({user.email})",
            exclude_admin=None
        )
        
        return {"status": "success", "message": "Password changed successfully"}
    except UserData.DoesNotExist:
        raise HTTPException(status_code=404, detail="User not found")

@router.post("/profile/image")
def upload_profile_image(file: UploadFile = File(...), admin: AdminUser = Depends(get_current_admin)):
    # Ensure database connection
    ensure_db_connection()
   
    # Note: This updates the ADMIN's profile image if AdminUser has that field.
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        MEDIA_ROOT = os.path.join(BASE_DIR, "media")
        PROFILE_PICS_DIR = os.path.join(MEDIA_ROOT, "profile_pics")
        os.makedirs(PROFILE_PICS_DIR, exist_ok=True)
       
        filename = f"admin_{admin.id}_{int(datetime.now().timestamp())}_{file.filename}"
        file_path = os.path.join(PROFILE_PICS_DIR, filename)
       
        with open(file_path, "wb+") as buffer:
            shutil.copyfileobj(file.file, buffer)
           
        full_image_url = f"{BASE_URL}/media/profile_pics/{filename}"
        return {"status": "success", "image_url": full_image_url}
       
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Image upload failed: {str(e)}")

@router.get("/profile/preferences")
def get_profile_preferences(admin: AdminUser = Depends(get_current_admin)):
    # Ensure database connection
    ensure_db_connection()
   
    return {"theme": "System", "time_zone": "UTC", "date_format": "ISO", "default_dashboard": "Overview"}

@router.put("/profile/preferences")
def update_profile_preferences(data: PreferencesSchema, admin: AdminUser = Depends(get_current_admin)):
    # Ensure database connection
    ensure_db_connection()
   
    return {"status": "success", "message": "Preferences updated"}

# ==============================================================================
# 📥 13. EXPORT USERS
# ==============================================================================
@router.get("/users/export")
def export_users_custom(
    format: str = Query("csv", enum=["csv", "excel"], description="File format"),
    from_date: Optional[date] = Query(None, description="Filter from YYYY-MM-DD"),
    to_date: Optional[date] = Query(None, description="Filter to YYYY-MM-DD"),
    role: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    admin: AdminUser = Depends(get_current_admin)
):
    # Ensure database connection
    ensure_db_connection()
   
    query = UserData.objects.all().order_by('-created_at')
    if role:
        query = query.filter(role__iexact=role)
    if status:
        query = query.filter(status__iexact=status)
    if search:
        query = query.filter(Q(full_name__icontains=search) | Q(email__icontains=search))
    if from_date:
        query = query.filter(created_at__date__gte=from_date)
    if to_date:
        query = query.filter(created_at__date__lte=to_date)
    users = list(query)
    
    # 🔔 NOTIFICATION: Admin exported users list
    export_count = len(users)
    create_admin_notification(
        admin=admin,
        notification_type="users_exported",
        title="Users Data Exported",
        subtitle=f"Admin {admin.name or admin.email} exported {export_count} users in {format.upper()} format"
    )
    
    headers = ["Full Name", "Email", "Role", "Status", "Joined Date"]
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
       
        for u in users:
            full_name = u.full_name or ""
            joined = u.created_at.strftime("%d-%b-%Y") if u.created_at else ""
            user_status = getattr(u, 'status', 'Active')
            writer.writerow([full_name, u.email, u.role, user_status, joined])
           
        output.seek(0)
        filename = f"users_export_{datetime.now().strftime('%Y%m%d')}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    elif format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users List"
        ws.append(headers)
       
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="2e1065", end_color="2e1065", fill_type="solid")
        for u in users:
            full_name = u.full_name or ""
            joined = u.created_at.strftime("%d-%b-%Y") if u.created_at else ""
            user_status = getattr(u, 'status', 'Active')
            ws.append([full_name, u.email, u.role, user_status, joined])
           
        for col in ws.columns:
             max_length = 0
             column = col[0].column_letter
             for cell in col:
                 try:
                     if len(str(cell.value)) > max_length:
                         max_length = len(str(cell.value))
                 except:
                     pass
             adjusted_width = (max_length + 2) * 1.2
             ws.column_dimensions[column].width = adjusted_width
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"users_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

# ==============================================================================
# 👤 14. ADMIN PROFILE MANAGEMENT
# ==============================================================================
@router.get("/profile")
async def get_admin_profile(admin: AdminUser = Depends(get_current_admin)):
    """Get admin profile information with S3 support"""
    ensure_db_connection()
   
    try:
        from asgiref.sync import sync_to_async
        
        # ✅ FIX: Use sync_to_async to get admin data
        admin_data = await sync_to_async(
            lambda: {
                "id": admin.id,
                "name": admin.name,
                "email": admin.email,
                "role": admin.role,
                "profile_image": admin.profile_image,
                "two_factor_enabled": getattr(admin, 'two_factor_enabled', False)
            }
        )()
        
        name_parts = admin_data['name'].split(' ', 1) if admin_data['name'] else ["", ""]
        first_name = name_parts[0] if len(name_parts) > 0 else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""
       
        # Get profile image URL with S3 support
        profile_image = None
        if admin_data['profile_image']:
            profile_image = get_admin_profile_image_url(str(admin_data['profile_image']))
       
        return {
            "id": admin_data['id'],
            "first_name": first_name,
            "last_name": last_name,
            "full_name": admin_data['name'],
            "email": admin_data['email'],
            "role": admin_data['role'],
            "two_factor_enabled": admin_data['two_factor_enabled'],
            "profile_image": profile_image,
            "storage_mode": "s3" if USE_S3 else "local"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/profile")
def update_admin_profile(
    data: AdminProfileUpdateSchema,
    admin: AdminUser = Depends(get_current_admin)
):
    """Update admin profile information"""
    ensure_db_connection()
   
    try:
        # print(f"🔔 DEBUG: update_admin_profile called for admin {admin.email}")
        # print(f"🔔 DEBUG: data received - first_name: {data.first_name}, last_name: {data.last_name}, email: {data.email}")
        
        old_name = admin.name
        old_email = admin.email
        old_role = admin.role
        
        # Update name...
        if data.first_name is not None or data.last_name is not None:
            current_first = data.first_name if data.first_name is not None else (admin.name.split(' ')[0] if admin.name else "")
            current_last = data.last_name if data.last_name is not None else (admin.name.split(' ')[1] if admin.name and len(admin.name.split(' ')) > 1 else "")
            admin.name = f"{current_first} {current_last}".strip()
       
        # Update email...
        if data.email and data.email != admin.email:
            if AdminUser.objects.filter(email=data.email).exclude(id=admin.id).exists():
                raise HTTPException(status_code=400, detail="Email already in use")
            admin.email = data.email
       
        # Update role...
        if data.role:
            admin.role = data.role
       
        admin.save()
        # print(f"✅ DEBUG: Admin saved successfully")
        
        # Check for changes
        changes = []
        if admin.name != old_name:
            changes.append(f"name changed to '{admin.name}'")
            # print(f"🔔 DEBUG: Name changed from '{old_name}' to '{admin.name}'")
        if admin.email != old_email:
            changes.append(f"email changed to '{admin.email}'")
            # print(f"🔔 DEBUG: Email changed from '{old_email}' to '{admin.email}'")
        if admin.role != old_role:
            changes.append(f"role changed to '{admin.role}'")
            # print(f"🔔 DEBUG: Role changed from '{old_role}' to '{admin.role}'")
        
        if changes:
            # print(f"🔔 DEBUG: Creating notification with changes: {changes}")
            result = create_notification_for_all_admins(
                notification_type="admin_profile_updated",
                title="Admin Profile Updated",
                subtitle=f"Admin {admin.name or admin.email} updated their profile: {', '.join(changes)}",
                exclude_admin=None
            )
            # print(f"🔔 DEBUG: Notification creation result: {result}")
        else:
            # print(f"🔔 DEBUG: No changes detected, skipping notification")
            pass
       
        return {
            "status": "success",
            "message": "Profile updated successfully",
            "user": {
                "id": admin.id,
                "first_name": data.first_name or (admin.name.split(' ')[0] if admin.name else ""),
                "last_name": data.last_name or (admin.name.split(' ')[1] if admin.name and len(admin.name.split(' ')) > 1 else ""),
                "email": admin.email,
                "role": admin.role
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        # print(f"❌ DEBUG: Error in update_admin_profile: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/profile/change-password")
def change_admin_password(
    data: AdminPasswordChangeSchema,
    admin: AdminUser = Depends(get_current_admin)
):
    """Change admin password"""
    # Ensure database connection
    ensure_db_connection()
   
    try:
        # Verify current password
        if not admin.check_password(data.current_password):
            raise HTTPException(status_code=400, detail="Current password is incorrect")
       
        # Set new password
        admin.set_password(data.new_password)
        admin.save()
        
        # 🔔 NOTIFICATION: Admin changed password
        create_notification_for_all_admins(
            notification_type="admin_password_changed",
            title="Admin Password Changed",
            subtitle=f"Admin {admin.name or admin.email} changed their password",
            exclude_admin=None
        )
       
        return {"status": "success", "message": "Password changed successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/profile/toggle-2fa")
def toggle_two_factor(
    enabled: bool = Query(..., description="Enable or disable 2FA"),
    admin: AdminUser = Depends(get_current_admin)
):
    """Toggle two-factor authentication"""
    # Ensure database connection
    ensure_db_connection()
   
    try:
        # If you need to add two_factor_enabled field to AdminUser model
        # Uncomment this after adding the field to models.py
        """
        admin.two_factor_enabled = enabled
        admin.save()
        """
        
        # 🔔 NOTIFICATION: Admin toggled 2FA
        create_notification_for_all_admins(
            notification_type="admin_2fa_toggled",
            title="2FA Setting Changed",
            subtitle=f"Admin {admin.name or admin.email} {'enabled' if enabled else 'disabled'} two-factor authentication",
            exclude_admin=None
        )
       
        return {
            "status": "success",
            "message": f"Two-factor authentication {'enabled' if enabled else 'disabled'}",
            "enabled": enabled
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/profile/upload-image")
async def upload_admin_profile_image(
    file: UploadFile = File(...),
    admin: AdminUser = Depends(get_current_admin)
):
    """
    Upload admin profile image with S3 support
    S3 Folder Used: admin_profiles/
    """
    ensure_db_connection()

    try:
        from asgiref.sync import sync_to_async
        
        # Save image using S3-aware function (this is async, so it's fine)
        saved_path = await save_admin_profile_image_s3(file, admin.id)
        
        # Delete old image if exists (async function)
        if admin.profile_image:
            await delete_admin_profile_image_s3(str(admin.profile_image))
        
        # ✅ FIX: Use sync_to_async for Django ORM operations
        await sync_to_async(
            lambda: setattr(admin, 'profile_image', saved_path)
        )()
        await sync_to_async(admin.save)()
        
        # Get the URL (synchronous function)
        image_url = get_admin_profile_image_url(saved_path)
        
        # ✅ FIX: Use sync_to_async for notification
        await sync_to_async(create_notification_for_all_admins)(
            notification_type="admin_image_uploaded",
            title="Admin Profile Image Updated",
            subtitle=f"Admin {admin.name or admin.email} updated their profile image",
            exclude_admin=None
        )
        
        return {
            "status": "success",
            "message": "Profile image uploaded successfully",
            "image_url": image_url,
            "storage_mode": "s3" if USE_S3 else "local"
        }

    except HTTPException:
        raise
    except Exception as e:
        # print("UPLOAD ERROR:", str(e))
        raise HTTPException(
            status_code=500,
            detail=f"Image upload failed: {str(e)}"
        )
        
@router.delete("/profile/remove-image")
async def remove_admin_profile_image(
    admin: AdminUser = Depends(get_current_admin)
):
    """Remove admin profile image with S3 support"""
    ensure_db_connection()
   
    try:
        from asgiref.sync import sync_to_async
        
        if admin.profile_image:
            # Delete from S3 or local (async function)
            await delete_admin_profile_image_s3(str(admin.profile_image))
            
            # ✅ FIX: Use sync_to_async for Django ORM operations
            await sync_to_async(
                lambda: setattr(admin, 'profile_image', None)
            )()
            await sync_to_async(admin.save)()
        
        # ✅ FIX: Use sync_to_async for notification
        await sync_to_async(create_notification_for_all_admins)(
            notification_type="admin_image_removed",
            title="Admin Profile Image Removed",
            subtitle=f"Admin {admin.name or admin.email} removed their profile image",
            exclude_admin=None
        )
       
        return {
            "status": "success",
            "message": "Profile image removed successfully",
            "storage_mode": "s3" if USE_S3 else "local"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/preferences")
def get_admin_preferences(admin: AdminUser = Depends(get_current_admin)):
    """Get admin preferences from database"""
    # Ensure database connection
    ensure_db_connection()
   
    try:
        # Default preferences
        default_preferences = {
            "theme": "Light",
            "time_zone": "UTC (Coordinated Universal Time)",
            "date_format": "ISO Format (YYYY-MM-DD)",
            "default_dashboard": "Overview Dashboard"
        }
       
        # Get preferences from admin user if they exist
        if hasattr(admin, 'preferences') and admin.preferences:
            # Merge with defaults to ensure all fields exist
            preferences = {**default_preferences, **admin.preferences}
        else:
            preferences = default_preferences
       
        return preferences
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/preferences")
def update_admin_preferences(
    data: PreferencesSchema,
    admin: AdminUser = Depends(get_current_admin)
):
    """Update admin preferences in database"""
    # Ensure database connection
    ensure_db_connection()
   
    try:
        # Convert to dict
        preferences = data.dict()
       
        # Save to admin user
        admin.preferences = preferences
        admin.save()
        
        # 🔔 NOTIFICATION: Admin updated preferences
        create_admin_notification(
            admin=admin,
            notification_type="admin_preferences_updated",
            title="Admin Preferences Updated",
            subtitle=f"Admin {admin.name or admin.email} updated their preferences"
        )
       
        return {
            "status": "success",
            "message": "Preferences updated successfully",
            "preferences": preferences
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==============================================================================
# ADMIN DELETE
# ==============================================================================
   
@router.delete("/profile")
def delete_admin_account(admin: AdminUser = Depends(get_current_admin)):
    ensure_db_connection()
    try:
        admin_name = admin.name
        admin_email = admin.email
        admin.delete()
        
        # 🔔 NOTIFICATION: Admin account deleted (notify other admins)
        create_notification_for_all_admins(
            notification_type="admin_account_deleted",
            title="Admin Account Deleted",
            subtitle=f"Admin account for {admin_name} ({admin_email}) has been deleted",
            exclude_admin=None
        )
        
        return {
            "status": "success",
            "message": "Admin account deleted successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
   
# ==============================================================================
# 🔔 15. ADMIN NOTIFICATIONS
# ==============================================================================
@router.get("/notifications")
def get_admin_notifications_api(admin: AdminUser = Depends(get_current_admin)):
    """Get all admin action notifications for the current admin"""
    ensure_db_connection()
    notifications = get_admin_notifications(admin)
    return notifications


@router.post("/notifications/{notification_id}/read")
def mark_admin_notification_read(
    notification_id: int,
    admin: AdminUser = Depends(get_current_admin)
):
    """Mark a single notification as read"""
    ensure_db_connection()
    
    updated = AdminNotification.objects.filter(
        id=notification_id, 
        admin=admin
    ).update(is_read=True)
    
    if updated:
        logger.info(f"📬 Marked notification {notification_id} as read for {admin.email}")
    
    return {
        "status": "success",
        "message": "Notification marked as read"
    }


@router.post("/notifications/mark-all-read")
def mark_all_admin_notifications_read(admin: AdminUser = Depends(get_current_admin)):
    """Mark all notifications as read"""
    ensure_db_connection()
    
    updated_count = AdminNotification.objects.filter(
        admin=admin, 
        is_read=False
    ).update(is_read=True)
    
    logger.info(f"📬 Marked {updated_count} notifications as read for {admin.email}")
    
    return {
        "status": "success",
        "message": f"{updated_count} notifications marked as read",
        "updated_count": updated_count
    }


@router.delete("/notifications/clear-all")
def clear_all_notifications_endpoint(admin: AdminUser = Depends(get_current_admin)):
    """
    Clear/delete ALL admin notifications for the current admin.
    Other admins' notifications are NOT affected.
    """
    ensure_db_connection()
    
    deleted_count = clear_all_notifications(admin)
    
    return {
        "status": "success",
        "message": f"Cleared {deleted_count} notifications",
        "deleted_count": deleted_count
    }